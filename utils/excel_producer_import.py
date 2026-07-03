"""
Парсер реального формата файлов, которые готовит аналитик DATFO по производителю
(пример: astra_zeneca-fom-II-Q-2026__май___.xlsx).

Формат файла (может отличаться в деталях от производителя к производителю,
поэтому колонки ищутся по заголовкам, а не по фиксированным индексам):

  • Лист "свод" (например "svod") — план/факт ПО ЮРЛИЦУ (ИНН), помесячно:
      № | ИНН | Официальное юридическое название | Аптеки | Кол-во аптек |
      ПЛАН-<квартал> | план-<месяц1> | факт-<месяц1> | ВП (%) | план-<месяц2> | ... |
      Регион | Район
    Важно: ИНН здесь — это сеть/юрлицо, а не конкретная аптечная точка.

  • Листы-месяцы (например "aprel", "may") — детализация ПО ТОЧКАМ:
      Код лекарства | Лекарство | FOM ID | ИНН | Официальное юридическое название |
      Наименование аптеки | Адрес | Область аптеки | Приход | [Внутренний приход] |
      Кол-во проданных товаров | Остатки | CIP | [Total zakub] | Total pradaja / total (UZS) |
      [Bonus] | [Total bonus]
    FOM ID — стабильный уникальный идентификатор конкретной физической аптеки
    (в отличие от ИНН, который у сети из 20+ точек один и тот же).

Результат парсинга — структура с отдельным периодом на каждый лист-месяц и
привязанной к нему сводкой план/факт по юрлицам (если удалось сопоставить месяц).
"""

import re
import openpyxl
from io import BytesIO

# Соответствие "название листа" (обычно латиницей/узб.) -> "суффикс месяца в свод-листе" (кириллица)
MONTH_SHEET_TO_SVOD_SUFFIX = {
    "yanvar": "январ", "fevral": "феврал", "mart": "март", "aprel": "апрел",
    "may": "май", "iyun": "июн", "iyul": "июл", "avgust": "август",
    "sentabr": "сентябр", "oktabr": "октябр", "noyabr": "ноябр", "dekabr": "декабр",
}
MONTH_DISPLAY_RU = {
    "yanvar": "Январь", "fevral": "Февраль", "mart": "Март", "aprel": "Апрель",
    "may": "Май", "iyun": "Июнь", "iyul": "Июль", "avgust": "Август",
    "sentabr": "Сентябрь", "oktabr": "Октябрь", "noyabr": "Ноябрь", "dekabr": "Декабрь",
}


def _find_col(headers: list[str], keywords: list[str], exclude: list[str] | None = None):
    """Find first header index containing any of keywords (case-insensitive), excluding some."""
    exclude = exclude or []
    for i, h in enumerate(headers):
        hl = h.lower()
        if any(ex in hl for ex in exclude):
            continue
        if any(kw in hl for kw in keywords):
            return i
    return None


def _header_row(ws, max_scan: int = 3):
    """Return (row_index, [str headers]) for the first row that looks like a header."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        cells = [str(v or "").strip() for v in row]
        joined = " ".join(cells).lower()
        if "fom id" in joined or ("инн" in joined and ("аптек" in joined or "лекарств" in joined)):
            return i, cells
    return None, None


def _is_svod_sheet(headers: list[str]) -> bool:
    joined = " ".join(headers).lower()
    return "инн" in joined and any(h.lower().startswith("план-") for h in headers)


def _is_detail_sheet(headers: list[str]) -> bool:
    joined = " ".join(headers).lower()
    return "fom id" in joined and "инн" in joined


def _safe_float(val) -> float:
    if val is None or val == "" or val == "-":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val) -> str:
    return str(val).strip() if val is not None else ""


def _extract_producer_name(svod_title: str, filename: str = "") -> str:
    """'Astra Zeneca-Fom-II-Q (май)' -> 'Astra Zeneca'."""
    if svod_title:
        m = re.split(r"[-_]?\s*fom\b", svod_title, flags=re.IGNORECASE)
        name = m[0].strip(" -_")
        if name:
            return name
    if filename:
        m = re.split(r"[-_]?\s*fom\b", filename, flags=re.IGNORECASE)
        name = m[0].replace("_", " ").strip(" -_")
        if name:
            return name.title()
    return "Неизвестный производитель"


def _parse_svod_sheet(ws) -> dict:
    """Returns {month_suffix: [entity rows]}."""
    header_idx, headers = _header_row(ws, max_scan=3)
    if not headers:
        return {}

    idx_inn    = _find_col(headers, ["инн"])
    idx_org    = _find_col(headers, ["юридическое"])
    idx_count  = _find_col(headers, ["кол-во аптек", "кол-во апт"])
    idx_region = _find_col(headers, ["регион"])
    idx_district = _find_col(headers, ["район"])

    if idx_inn is None:
        return {}

    # Найти пары (план-<мес>, факт-<мес>) по заголовкам
    month_cols = {}  # suffix -> (plan_idx, fact_idx)
    for i, h in enumerate(headers):
        hl = h.lower()
        if hl.startswith("план-"):
            suffix = hl.split("план-", 1)[1].strip()
            fact_idx = None
            for j in range(i + 1, len(headers)):
                if headers[j].lower().startswith(f"факт-{suffix}"):
                    fact_idx = j
                    break
            if fact_idx is not None:
                month_cols[suffix] = (i, fact_idx)

    result: dict[str, list] = {suffix: [] for suffix in month_cols}

    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        try:
            inn = _safe_str(row[idx_inn]) if idx_inn < len(row) else ""
            if not inn or inn == "-" or not inn.replace(".", "").isdigit():
                continue  # пропускаем строку итогов и пустые строки

            org_name = _safe_str(row[idx_org]) if idx_org is not None and idx_org < len(row) else ""
            count    = int(_safe_float(row[idx_count])) if idx_count is not None and idx_count < len(row) else 0
            region   = _safe_str(row[idx_region]) if idx_region is not None and idx_region < len(row) else ""
            district = _safe_str(row[idx_district]) if idx_district is not None and idx_district < len(row) else ""

            for suffix, (plan_idx, fact_idx) in month_cols.items():
                plan_amt = _safe_float(row[plan_idx]) if plan_idx < len(row) else 0.0
                fact_amt = _safe_float(row[fact_idx]) if fact_idx < len(row) else 0.0
                result[suffix].append({
                    "inn": inn, "org_name": org_name, "pharmacy_count": count,
                    "region": region, "district": district,
                    "plan_amount": plan_amt, "fact_amount": fact_amt,
                })
        except Exception as e:
            print(f"_parse_svod_sheet: skipping malformed row: {e}")
            continue

    return result


def _parse_detail_sheet(ws) -> list[dict]:
    header_idx, headers = _header_row(ws, max_scan=2)
    if not headers:
        return []

    idx_sku      = _find_col(headers, ["код лекарства", "код sku", "sku"])
    idx_product  = _find_col(headers, ["лекарств", "продукт"], exclude=["код"])
    idx_fom      = _find_col(headers, ["fom id"])
    idx_inn      = _find_col(headers, ["инн"])
    idx_org      = _find_col(headers, ["юридическое"])
    idx_pharmacy = _find_col(headers, ["наименование аптеки"]) or _find_col(headers, ["аптек"], exclude=["область", "кол-во"])
    idx_address  = _find_col(headers, ["адрес"])
    idx_region   = _find_col(headers, ["область", "регион"])
    idx_incoming = _find_col(headers, ["приход"], exclude=["внутренний"])
    idx_sales_qty= _find_col(headers, ["кол-во проданных", "продажи шт", "продано"])
    idx_stock    = _find_col(headers, ["остат"])
    idx_price    = _find_col(headers, ["cip"])
    idx_sales_amt= _find_col(headers, ["pradaja"]) or _find_col(headers, ["total"], exclude=["zakub", "bonus"])
    idx_bonus_rate = _find_col(headers, ["bonus"], exclude=["total"])
    idx_total_bonus = _find_col(headers, ["total bonus"])

    if idx_fom is None or idx_inn is None or idx_product is None:
        return []

    def get(row, idx, default=None):
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    rows = []
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        try:
            fom_id = get(row, idx_fom)
            if fom_id in (None, ""):
                continue
            product = _safe_str(get(row, idx_product))
            if not product:
                continue

            rows.append({
                "fom_id":       str(int(fom_id)) if str(fom_id).replace(".", "").isdigit() else str(fom_id).strip(),
                "sku_code":     _safe_str(get(row, idx_sku)) or product[:20],
                "product":      product,
                "inn":          _safe_str(get(row, idx_inn)),
                "org_name":     _safe_str(get(row, idx_org)),
                "pharmacy":     _safe_str(get(row, idx_pharmacy)) or _safe_str(get(row, idx_org)),
                "address":      _safe_str(get(row, idx_address)),
                "region":       _safe_str(get(row, idx_region)),
                "incoming_qty": _safe_float(get(row, idx_incoming, 0)),
                "sales_qty":    _safe_float(get(row, idx_sales_qty, 0)),
                "stock_qty":    _safe_float(get(row, idx_stock, 0)),
                "unit_price":   _safe_float(get(row, idx_price, 0)),
                "sales_amount": _safe_float(get(row, idx_sales_amt, 0)),
                "bonus_rate":   _safe_float(get(row, idx_bonus_rate, 0)),
                "total_bonus":  _safe_float(get(row, idx_total_bonus, 0)),
            })
        except Exception as e:
            print(f"_parse_detail_sheet: skipping malformed row: {e}")
            continue

    return rows


def parse_producer_workbook(file_bytes: bytes, filename: str = "") -> dict | None:
    """
    Парсит файл в формате DATFO (свод + листы-месяцы).
    Возвращает None если формат не распознан (тогда стоит попробовать старый
    плоский парсер parse_analytics_excel как запасной вариант).

    Result:
    {
        "producer_name": str,
        "periods": [
            {
                "period_key": "may",
                "period_name": "Май 2026",
                "rows": [ ...detail rows... ],
                "entity_plans": [ ...entity-level plan/fact rows... ],
            },
            ...
        ],
        "warnings": [str, ...],
    }
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        print(f"parse_producer_workbook: cannot open workbook: {e}")
        return None

    svod_entity_plans = {}
    svod_title = ""
    detail_sheets = []  # (sheet_name, ws)
    warnings = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_idx, headers = _header_row(ws, max_scan=3)
        if not headers:
            continue
        if _is_svod_sheet(headers):
            svod_entity_plans = _parse_svod_sheet(ws)
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row and first_row[0]:
                svod_title = str(first_row[0])
        elif _is_detail_sheet(headers):
            detail_sheets.append((sheet_name, ws))

    if not detail_sheets:
        return None  # не похоже на этот формат вообще

    producer_name = _extract_producer_name(svod_title, filename)

    # Определяем год для отображаемого имени периода
    year_match = re.search(r"(20\d{2})", filename) or re.search(r"(20\d{2})", svod_title)
    year = year_match.group(1) if year_match else ""

    periods = []
    for sheet_name, ws in detail_sheets:
        rows = _parse_detail_sheet(ws)
        if not rows:
            warnings.append(f"Лист «{sheet_name}»: не удалось распознать колонки, пропущен.")
            continue

        key = sheet_name.strip().lower()
        display_month = MONTH_DISPLAY_RU.get(key, sheet_name.strip().title())
        period_name = f"{display_month} {year}".strip()

        svod_suffix = MONTH_SHEET_TO_SVOD_SUFFIX.get(key, key)
        entity_plans = svod_entity_plans.get(svod_suffix, [])
        if not entity_plans:
            warnings.append(f"Лист «{sheet_name}»: план/факт по юрлицам из «свод» не найден для этого месяца.")

        periods.append({
            "period_key": key,
            "period_name": period_name,
            "rows": rows,
            "entity_plans": entity_plans,
        })

    if not periods:
        return None

    return {
        "producer_name": producer_name,
        "periods": periods,
        "warnings": warnings,
    }

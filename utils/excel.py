"""
Excel parser for DATFO analytics bot.

Expected format (sheet "Data" or first sheet):
Row 1: Headers
  A: Производитель (producer name)
  B: Код SKU (sku_code)
  C: Продукт (product name)
  D: ИНН аптеки
  E: Аптека (pharmacy name)
  F: Регион
  G: Город
  H: Адрес
  I: Остаток (stock qty)
  J: Продажи шт (sales qty)
  K: Продажи сум (sales amount)
  L: План шт (plan qty)
  M: План сум (plan amount)
  N: Факт шт (fact qty)
  O: Факт сум (fact amount)
  P: Бонус (bonus amount)
  Q: Дата (date)
Row 2+: Data rows

Sheet "Период" (optional):
  A1: название периода
"""

import openpyxl
from io import BytesIO
from datetime import date


def parse_analytics_excel(file_bytes: bytes) -> dict | None:
    """
    Parse analytics Excel file.
    Returns structured data for import.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

        # Get period name
        period_name = None
        if "Период" in wb.sheetnames:
            ws_p = wb["Период"]
            period_name = str(ws_p.cell(1, 1).value or "").strip()

        # Get data sheet
        ws = None
        for sheet_name in ["Data", "Данные", "Sheet1"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active

        if not period_name:
            # Try to get from first sheet cell A1 if it looks like a period
            val = ws.cell(1, 1).value
            if val and str(val).strip():
                period_name = str(val).strip()

        rows = []
        header_row = None

        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if all(v is None for v in row):
                continue
            # Find header row
            if header_row is None:
                row_str = [str(v or "").lower() for v in row]
                if any(k in " ".join(row_str) for k in ["производитель", "аптека", "продукт", "sku"]):
                    header_row = i
                    continue
                elif i > 5:
                    # Assume row 1 is header
                    header_row = 1

            if header_row and i > header_row:
                producer  = str(row[0] or "").strip()
                sku_code  = str(row[1] or "").strip()
                product   = str(row[2] or "").strip()
                inn       = str(row[3] or "").strip()
                pharmacy  = str(row[4] or "").strip()
                region    = str(row[5] or "").strip()
                city      = str(row[6] or "").strip()
                address   = str(row[7] or "").strip()

                if not producer or not inn or not product:
                    continue

                def safe_float(val):
                    try:
                        return float(val) if val not in (None, "", "-") else 0.0
                    except (ValueError, TypeError):
                        return 0.0

                stock_qty    = safe_float(row[8])
                sales_qty    = safe_float(row[9])
                sales_amount = safe_float(row[10])
                plan_qty     = safe_float(row[11])
                plan_amount  = safe_float(row[12])
                fact_qty     = safe_float(row[13])
                fact_amount  = safe_float(row[14])
                bonus        = safe_float(row[15])

                date_val = row[16] if len(row) > 16 else None
                if hasattr(date_val, 'date'):
                    date_val = date_val.date()
                elif isinstance(date_val, str):
                    try:
                        from datetime import datetime
                        date_val = datetime.strptime(date_val[:10], "%Y-%m-%d").date()
                    except Exception:
                        date_val = date.today()
                else:
                    date_val = date.today()

                # Try to get INN as integer string
                try:
                    inn = str(int(float(inn)))
                except (ValueError, TypeError):
                    pass

                rows.append({
                    "producer":     producer,
                    "sku_code":     sku_code or product[:20],
                    "product":      product,
                    "inn":          inn,
                    "pharmacy":     pharmacy,
                    "region":       region,
                    "city":         city,
                    "address":      address,
                    "stock_qty":    stock_qty,
                    "sales_qty":    sales_qty,
                    "sales_amount": sales_amount,
                    "plan_qty":     plan_qty,
                    "plan_amount":  plan_amount,
                    "fact_qty":     fact_qty,
                    "fact_amount":  fact_amount,
                    "bonus":        bonus,
                    "date_on":      date_val,
                })

        if not rows:
            return None

        # Count stats
        producers = set(r["producer"] for r in rows)
        pharmacies = set(r["inn"] for r in rows)
        products = set(r["product"] for r in rows)

        return {
            "period_name":    period_name or f"Обновление {date.today().strftime('%d.%m.%Y')}",
            "rows":           rows,
            "total_rows":     len(rows),
            "producers":      len(producers),
            "pharmacies":     len(pharmacies),
            "products":       len(products),
        }

    except Exception as e:
        print(f"parse_analytics_excel error: {e}")
        import traceback
        traceback.print_exc()
        return None


def fmt_money(amount: float) -> str:
    if not amount:
        return "0 сум"
    return f"{int(amount):,}".replace(",", " ") + " сум"


def fmt_qty(qty: float) -> str:
    if not qty:
        return "0"
    return f"{int(qty):,}".replace(",", " ")


def fmt_pct(fact: float, plan: float) -> str:
    if not plan:
        return "—"
    pct = round(fact / plan * 100, 1)
    emoji = "🟢" if pct >= 100 else "🟡" if pct >= 70 else "🔴"
    return f"{emoji} {pct}%"


def make_progress_bar(fact: float, plan: float) -> str:
    if not plan:
        return "░░░░░░░░░░ 0%"
    pct = min(fact / plan, 1.0)
    filled = round(pct * 10)
    bar = "█" * filled + "░" * (10 - filled)
    return f"`{bar}` {round(pct * 100)}%"
